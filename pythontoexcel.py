import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1, title='Second Sheet')


ws['A1'] = 'Invoice'


Fontobject = Font(name='Times New Roman', size=24, italic=False,bold=True)
ws['A1'].font = Fontobject


ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'


ws.merge_cells('A1:B1')
#unmerge cells
#ws.unmerge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16, bold=True)

ws['B8'] = '=SUM(B2:B4)'

ws.column_dimensions['A'].width = 25

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

for currentrow in read_ws.iter_rows(min_row=1,max_row=read_ws.max_row,max_col=read_ws.max_column):
    for currentcell in currentrow:
        write_sheet[currentcell.coordinate] = currentcell.value

maxrow = write_sheet.max_row

write_sheet[f'C{maxrow+3}'] = f'=SUM(C2:C{maxrow})'
write_sheet[f'D{maxrow+3}'] = f'=SUM(D2:D{maxrow})'
write_sheet[f'A{maxrow+3}'] = 'Grand Totals'

write_sheet[f'C{maxrow+4}'] = f'=AVERAGE(C2:C{maxrow})'
write_sheet[f'D{maxrow+4}'] = f'=AVERAGE(D2:D{maxrow})'
write_sheet[f'A{maxrow+4}'] = 'Average'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 16
write_sheet.column_dimensions['C'].width = 16
write_sheet.column_dimensions['D'].width = 16

for cell in write_sheet['C:C']:
    cell.number_format = '#,##0'

for cell in write_sheet['D:D']:
    cell.number_format = u'"$"#,##0.00'

wb.save('PythontoExcel.xlsx')