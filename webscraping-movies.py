
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'No.'


Fontobject = Font(name='Calibri', size=16, italic=False,bold=True)
ws['A1'].font = Fontobject
ws['B1'].font = Fontobject
ws['C1'].font = Fontobject
ws['D1'].font = Fontobject
ws['E1'].font = Fontobject
ws['F1'].font = Fontobject

ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Total Gross'
ws['E1'] = 'Theaters'
ws['F1'] = 'Average per theater'

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 25

#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

table_rows = soup.findAll('tr')

x=1
for row in table_rows[0:5]:
    ws[f'A{x+1}'] = str(x)
    title = table_rows[x].find(attrs={'class':'a-text-left mojo-field-type-release mojo-cell-wide'}).text
    ws[f'B{x+1}'] = title
    print(title)
    release = table_rows[x].find(attrs={'class':'a-text-left mojo-field-type-date a-nowrap'}).text
    ws[f'C{x+1}'] = release
    print(release)
    total_gross = int(table_rows[x].find(attrs={'class':'a-text-right mojo-field-type-money mojo-estimatable'}).text.strip('$').replace(',',''))
    ws[f'D{x+1}'] = total_gross
    print(total_gross)
    theaters = int(table_rows[x].find(attrs={'class':'a-text-right mojo-field-type-positive_integer'}).text.replace(',',''))
    ws[f'E{x+1}'] = theaters
    print(theaters)
    average_per_theater = total_gross/theaters
    ws[f'F{x+1}'] = average_per_theater
    print(average_per_theater)
    x+=1

for cell in ws['E:E']:
    cell.number_format = '#,##0'

for cell in ws['D:D']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['F:F']:
    cell.number_format = u'"$ "#,##0.00'

wb.save('webscraping_movies.xlsx')